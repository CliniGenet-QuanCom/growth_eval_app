# -*- coding: utf-8 -*-
"""Generate Dart const data tables from the three reference workbooks.
Output goes to <project>/lib/data/*.dart . The xlsx files themselves are
NOT redistributed; only the numeric coefficients are embedded."""
import openpyxl, os, io

DL = r"C:\Users\yseki\Downloads"
PROJ = r"C:\Users\yseki\Documents\Google Drive\\u4f53格指数の計算アプリ(Claude)"
# build path explicitly to avoid escape issues
PROJ = os.path.join(r"C:\Users\yseki\Documents\Google Drive", "体格指数の計算アプリ(Claude)")
OUT = os.path.join(PROJ, "lib", "data")
os.makedirs(OUT, exist_ok=True)

def wb(fn):
    return openpyxl.load_workbook(os.path.join(DL, fn), data_only=True)

def num(v):
    if v is None:
        return 0.0
    return float(v)

def fnum(v):
    # format a float for dart
    s = repr(float(v))
    return s

def write(name, content):
    p = os.path.join(OUT, name)
    with io.open(p, "w", encoding="utf-8") as f:
        f.write(content)
    print("wrote", p, len(content), "bytes")

HEADER = "// GENERATED FILE - do not edit by hand.\n// Coefficients transcribed from the JSPE/JSGA reference workbooks.\n\n"

# ---------- Height reference (2000) ----------
def gen_height():
    w = wb("taikakushisu_v3.3.xlsx")
    ws = w["Height"]
    def block(r0, c0):
        # 12 months x 18 years
        rows = []
        for m in range(12):
            row = [num(ws.cell(r0 + m, c0 + y).value) for y in range(18)]
            rows.append(row)
        return rows
    # Hmalemean B5:S16 -> r0=5,c0=2 ; Hmalesd U5:AL16 -> c0=21
    male_mean = block(5, 2)
    male_sd = block(5, 21)
    female_mean = block(22, 2)
    female_sd = block(22, 21)
    def fmt(name, data):
        lines = ["const List<List<double>> %s = [" % name]
        for row in data:
            lines.append("  [" + ", ".join(fnum(v) for v in row) + "],")
        lines.append("];")
        return "\n".join(lines)
    out = HEADER
    out += "// Height mean/SD (cm) indexed [completedMonth 0-11][ageYear 0-17]. 2000 national reference.\n"
    out += fmt("heightMeanMale", male_mean) + "\n\n"
    out += fmt("heightSdMale", male_sd) + "\n\n"
    out += fmt("heightMeanFemale", female_mean) + "\n\n"
    out += fmt("heightSdFemale", female_sd) + "\n"
    write("height_reference.dart", out)

# ---------- BMI LMS (cubic in decimal months) ----------
def gen_bmi():
    w = wb("taikakushisu_v3.3.xlsx")
    ws = w["BMILMS"]
    def coef(r):
        return [num(ws.cell(r, c).value) for c in (4, 5, 6, 7)]  # D,E,F,G = x3,x2,x1,x0
    # segments: list of (upperBoundMonths, [a,b,c,d]); last seg bound = 1e9
    male_L = [(78, coef(5)), (150, coef(6)), (1e9, coef(7))]
    female_L = [(69, coef(9)), (150, coef(10)), (1e9, coef(11))]
    male_M = [(2.5, coef(21)), (9.5, coef(22)), (26.75, coef(23)), (90, coef(24)), (1e9, coef(25))]
    female_M = [(2.5, coef(27)), (9.5, coef(28)), (26.75, coef(29)), (90, coef(30)), (150, coef(31)), (1e9, coef(32))]
    male_S = [(90, coef(14)), (1e9, coef(15))]
    female_S = [(90, coef(17)), (1e9, coef(18))]
    def fmt(name, segs):
        lines = ["const List<List<double>> %s = [" % name]
        for up, c in segs:
            lines.append("  [%s, %s, %s, %s, %s]," % (fnum(up), fnum(c[0]), fnum(c[1]), fnum(c[2]), fnum(c[3])))
        lines.append("];")
        return "\n".join(lines)
    out = HEADER
    out += "// BMI LMS cubic segments. Each row: [upperBoundMonths, a, b, c, d] for value = a*x^3+b*x^2+c*x+d, x = decimal months.\n"
    for n, s in [("bmiLMale", male_L), ("bmiLFemale", female_L), ("bmiMMale", male_M),
                 ("bmiMFemale", female_M), ("bmiSMale", male_S), ("bmiSFemale", female_S)]:
        out += fmt(n, s) + "\n\n"
    write("bmi_lms.dart", out)

# ---------- Weight SDS piecewise polynomial coeffs ----------
def gen_weight():
    w = wb("taikakushisu_v3.3.xlsx")
    ws = w["WeightSDS"]
    # columns M..W = 13..23 = powers 10..0
    def poly(r):
        return [num(ws.cell(r, c).value) for c in range(13, 24)]
    def fmtpoly(name, r):
        c = poly(r)
        return "const List<double> %s = [%s];" % (name, ", ".join(fnum(v) for v in c))
    out = HEADER
    out += "// Weight-SDS LMS: L,M,S are piecewise polynomials of integer age in months.\n"
    out += "// Polynomial coefficient lists are ordered power 10 -> 0 (11 entries).\n\n"
    # L
    out += fmtpoly("wL_male", 5) + "\n"
    out += fmtpoly("wL_female_lo", 8) + "  // months < 186\n"
    out += "// female L, months>=186: U9 + V9*(x-W9)\n"
    out += "const double wL_female_hi_U = %s;\n" % fnum(num(ws.cell(9, 21).value))
    out += "const double wL_female_hi_V = %s;\n" % fnum(num(ws.cell(9, 22).value))
    out += "const double wL_female_hi_W = %s;\n\n" % fnum(num(ws.cell(9, 23).value))
    # M
    out += fmtpoly("wM_male_a", 23) + "  // months < 45\n"
    out += fmtpoly("wM_male_b", 25) + "  // months < 153\n"
    out += "// male M, months>=153: M27 + N27/(1+exp(O27+P27*x))\n"
    out += "const List<double> wM_male_logit = [%s, %s, %s, %s];\n" % tuple(
        fnum(num(ws.cell(27, c).value)) for c in (13, 14, 15, 16))
    out += fmtpoly("wM_female_a", 29) + "  // months < 43.8 ; then -0.010431*(1-x/210)\n"
    out += fmtpoly("wM_female_b", 30) + "  // months < 123 ; then -0.010431*(1-1/x)\n"
    out += "// female M, months>=123: M32 + N32/(1+exp(O32+P32*x)) - 0.010431*(1-x/210)\n"
    out += "const List<double> wM_female_logit = [%s, %s, %s, %s];\n\n" % tuple(
        fnum(num(ws.cell(32, c).value)) for c in (13, 14, 15, 16))
    # S
    out += fmtpoly("wS_male_lo", 12) + "  // months < 162\n"
    out += fmtpoly("wS_male_hi", 14) + "  // months >= 162\n"
    out += fmtpoly("wS_female_lo", 17) + "  // months < 156\n"
    out += "// female S, 156<=months<186: U + (V-U)/24*(x-186) + Wc*(186-x)^2 - 0.005\n"
    out += "//            months>=186: U + (V-U)/24*(x-186) - 0.005\n"
    out += "const double wS_female_U = %s;\n" % fnum(num(ws.cell(18, 21).value))
    out += "const double wS_female_V = %s;\n" % fnum(num(ws.cell(18, 22).value))
    out += "const double wS_female_W = %s;\n" % fnum(num(ws.cell(18, 23).value))
    write("weight_sds.dart", out)

# ---------- StdBW (Murata / Ito) ----------
def gen_stdbw():
    w = wb("taikakushisu_v3.3.xlsx")
    ws = w["StdBW"]
    # itoOI H6:K11 (cols 8..11), 6 rows: male bands 3, female bands 3
    ito = []
    for r in range(6, 12):
        ito.append([num(ws.cell(r, c).value) for c in (8, 9, 10, 11)])
    # age-linear: muratamale C4:D16 (cols 3,4) ages 5..17 ; muratafemale C18:D30
    def agelin(r0):
        rows = []
        for i in range(13):
            rows.append([num(ws.cell(r0 + i, 3).value), num(ws.cell(r0 + i, 4).value)])  # a,b
        return rows
    male_age = agelin(4)
    female_age = agelin(18)
    out = HEADER
    out += "// Standard-weight coefficient tables for the three obesity indices.\n\n"
    out += "// Murata infant quadratic: stdWeight(kg) = a*h^2 + b*h + c, h in cm.\n"
    out += "const List<double> murataMale = [2.06e-3, -0.1166, 6.5273];\n"
    out += "const List<double> murataFemale = [2.49e-3, -0.1858, 9.036];\n\n"
    out += "// Ito height-band cubic: stdWeight(kg) = a*X^3+b*X^2+c*X+d, X = height(cm)/100.\n"
    out += "// Rows 0-2 male (h<140, 140-149, >=149), rows 3-5 female.\n"
    out += "const List<List<double>> itoHeightCubic = [\n"
    for row in ito:
        out += "  [%s, %s, %s, %s],\n" % tuple(fnum(v) for v in row)
    out += "];\n\n"
    out += "// Ito age-linear: stdWeight(kg) = a*h + b, h in cm. Index = INT(age)-5 for ages 5..17.\n"
    out += "const List<List<double>> itoAgeLinearMale = [\n"
    for row in male_age:
        out += "  [%s, %s],\n" % (fnum(row[0]), fnum(row[1]))
    out += "];\n"
    out += "const List<List<double>> itoAgeLinearFemale = [\n"
    for row in female_age:
        out += "  [%s, %s],\n" % (fnum(row[0]), fnum(row[1]))
    out += "];\n"
    write("stdbw.dart", out)

# ---------- IGF-I reference (annual L,M,S, ages 0..77) ----------
def gen_igf():
    w = wb("taikakushisu_v3.3.xlsx")
    ws = w["IGF-LMS"]
    # IGFmale C6:E83 (cols 3,4,5), IGFfemale N6:P83 (cols 14,15,16)
    def block(c0):
        rows = []
        for i in range(78):
            rows.append([num(ws.cell(6 + i, c0).value), num(ws.cell(6 + i, c0 + 1).value), num(ws.cell(6 + i, c0 + 2).value)])
        return rows
    male = block(3)
    female = block(14)
    out = HEADER
    out += "// IGF-I LMS by completed age in YEARS (0..77). Each row [L, M, S].\n"
    def fmt(name, data):
        s = "const List<List<double>> %s = [\n" % name
        for row in data:
            s += "  [%s, %s, %s],\n" % tuple(fnum(v) for v in row)
        s += "];\n"
        return s
    out += fmt("igfMale", male) + "\n" + fmt("igfFemale", female)
    write("igf_reference.dart", out)

# ---------- Birth-size reference (week 22..41 x day 0..6) ----------
def gen_birth():
    w = wb("taikakubirthlongcross_v1.1.xlsx")
    ws = w["新生児reference"]
    # blocks start columns (L at): maleFB D=4, maleSB J=10, femaleFB P=16, femaleSB V=22, head AB=28, birthH AH=34
    # range rows 7..146 (140 rows)
    def block(c0):
        rows = []
        for i in range(140):
            r = 7 + i
            rows.append([num(ws.cell(r, c0).value), num(ws.cell(r, c0 + 1).value), num(ws.cell(r, c0 + 2).value)])
        return rows
    tables = {
        "birthWeightMalePrimi": 4,
        "birthWeightMaleMulti": 10,
        "birthWeightFemalePrimi": 16,
        "birthWeightFemaleMulti": 22,
        "birthHead": 28,
        "birthLength": 34,
    }
    out = HEADER
    out += "// Gestational birth-size LMS. Row index = (week-22)*7 + day, weeks 22..41, days 0..6.\n"
    out += "// Weight tables in GRAMS; length & head in cm. Weight is sex x parity; length/head combined-sex.\n\n"
    for name, c0 in tables.items():
        data = block(c0)
        out += "const List<List<double>> %s = [\n" % name
        for row in data:
            out += "  [%s, %s, %s],\n" % tuple(fnum(v) for v in row)
        out += "];\n\n"
    write("birth_reference.dart", out)

gen_height()
gen_bmi()
gen_weight()
gen_stdbw()
gen_igf()
gen_birth()
print("DONE")
